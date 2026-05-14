
import csv, os, sys, time
from dataclasses import dataclass
from typing import Optional
from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import *
try: import pyqtgraph as pg
except Exception: pg=None
try: from pymodbus.client import ModbusTcpClient, ModbusSerialClient
except Exception: ModbusTcpClient=ModbusSerialClient=None
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook=load_workbook=None
CATEGORIES=['Drive Identification','Communication','Command and Reference','Monitoring','Faults and Diagnostics','Motor Setup','Application Functions','Protection and Limits','Ramp and Motion Profile','Input Output Configuration','Maintenance and Service']
def resource_path(p): return os.path.join(getattr(sys,'_MEIPASS',os.path.abspath(os.path.join(os.path.dirname(__file__),'..'))),p)
def sf(v,d=0.0):
    try: return d if v is None or str(v).strip()=='' else float(v)
    except: return d
def si(v,d=0):
    try: return int(float(v))
    except: return d
@dataclass
class P:
    code:str; name:str; address:int; datatype:str; scale:float; default:float; min:float; max:float; unit:str; access:str; monitor:bool; wp:bool; scope:bool; group:str; sub:str; policy:str; value:Optional[float]=None; online:Optional[float]=None; mod:bool=False
    @property
    def eff(self): return self.value if self.value is not None else self.default
class DB:
    def __init__(self): self.params=[]
    def load(self,path):
        self.params=[]
        with open(path,newline='',encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                self.params.append(P(r.get('code',''),r.get('name',''),si(r.get('address')),r.get('datatype','uint16'),sf(r.get('scale'),1),sf(r.get('default')),sf(r.get('min')),sf(r.get('max'),65535),r.get('unit',''),r.get('access','RO'),str(r.get('monitor','')).upper()=='TRUE',str(r.get('write_protect','')).upper()=='TRUE',str(r.get('scope','')).upper()=='TRUE',r.get('group','Monitoring'),r.get('subcategory',''),r.get('write_policy','')))
    def filt(self,s='',g='All',mon=False):
        s=(s or '').lower().strip(); exact=s.startswith('='); s=s[1:].strip() if exact else s; out=[]
        for p in self.params:
            if mon and not p.monitor: continue
            if g!='All' and p.group!=g: continue
            if s:
                if exact and p.code.lower()!=s: continue
                if not exact and not (s in p.code.lower() or s in p.name.lower() or s in str(p.address) or s in p.group.lower()): continue
            out.append(p)
        return out
    def by(self,code):
        for p in self.params:
            if p.code.upper()==code.upper(): return p
class GW:
    def __init__(self): self.client=None; self.unit=1; self.off=0; self.mode='TCP'
    def connect_tcp(self,h,port,unit,zero):
        if ModbusTcpClient is None: raise RuntimeError('pymodbus not installed')
        self.mode='TCP'; self.unit=unit; self.off=-1 if zero else 0; self.client=ModbusTcpClient(host=h,port=port,timeout=3)
        if not self.client.connect(): raise RuntimeError('Could not connect TCP')
    def connect_rtu(self,port,baud,parity,stop,bits,unit,zero):
        if ModbusSerialClient is None: raise RuntimeError('pymodbus/pyserial not installed')
        self.mode='RTU'; self.unit=unit; self.off=-1 if zero else 0
        try:
            self.client=ModbusSerialClient(method='rtu',port=port,baudrate=baud,parity=parity,stopbits=stop,bytesize=bits,timeout=2)
        except TypeError:
            self.client=ModbusSerialClient(port=port,baudrate=baud,parity=parity,stopbits=stop,bytesize=bits,timeout=2)
        if not self.client.connect(): raise RuntimeError('Could not open RTU serial port')
    def close(self):
        if self.client: self.client.close()
        self.client=None
    def ok(self): return self.client is not None
    def kw(self): return [{'slave':self.unit},{'unit':self.unit},{'device_id':self.unit},{}]
    def rr(self,a,c=1):
        last=None
        for k in self.kw():
            try:
                r=self.client.read_holding_registers(address=a+self.off,count=c,**k)
                if r.isError(): raise RuntimeError(str(r))
                return r.registers
            except TypeError as e: last=e
        raise RuntimeError(last)
    def wr(self,a,v):
        last=None
        for k in self.kw():
            try:
                r=self.client.write_register(address=a+self.off,value=int(v)&0xffff,**k)
                if r.isError(): raise RuntimeError(str(r))
                return
            except TypeError as e: last=e
        raise RuntimeError(last)
    def readp(self,p):
        regs=self.rr(p.address,2 if p.datatype in ('uint32','int32') else 1)
        if p.datatype=='int16': raw=regs[0] if regs[0]<32768 else regs[0]-65536
        elif p.datatype=='uint32': raw=(regs[0]<<16)+regs[1]
        elif p.datatype=='int32':
            raw=(regs[0]<<16)+regs[1]; raw=raw-4294967296 if raw>=2147483648 else raw
        else: raw=regs[0]
        return raw*p.scale
    def writep(self,p,val):
        raw=int(round(val/(p.scale or 1)))
        if p.datatype=='int16' and raw<0: raw=65536+raw
        self.wr(p.address,raw)
class W(QMainWindow):
    def __init__(self):
        super().__init__(); self.setWindowTitle('LK XD4000 Phase-5E TCP/RTU + Oscilloscope'); self.resize(1580,930)
        self.db=DB(); self.gw=GW(); self.params=[]; self.scope_data={}; self.scope_start=None
        self.scope_timer=QTimer(self); self.scope_timer.timeout.connect(self.scope_tick)
        self.keep_timer=QTimer(self); self.keep_timer.timeout.connect(self.keep_tick)
        self.ui(); self.db.load(resource_path('data/xd4000_phase5e_full_parameters.csv')); self.refresh(); self.log(f'Loaded parameter database: {len(self.db.params)} parameters')
    def ui(self):
        c=QWidget(); self.setCentralWidget(c); root=QVBoxLayout(c)
        box=QGroupBox('Connection - Modbus TCP / Modbus RTU RS485'); g=QGridLayout(box)
        self.mode=QComboBox(); self.mode.addItems(['TCP','RTU-RS485'])
        self.host=QLineEdit('192.168.1.10'); self.tcp_port=QSpinBox(); self.tcp_port.setRange(1,65535); self.tcp_port.setValue(502)
        self.com=QLineEdit('COM3'); self.baud=QComboBox(); self.baud.addItems(['9600','19200','38400','57600','115200']); self.baud.setCurrentText('19200')
        self.parity=QComboBox(); self.parity.addItems(['E','N','O']); self.stop=QComboBox(); self.stop.addItems(['1','2']); self.bits=QComboBox(); self.bits.addItems(['8','7'])
        self.unit=QSpinBox(); self.unit.setRange(1,255); self.unit.setValue(1); self.zero=QCheckBox('Zero-based (-1)')
        for i,(l,w) in enumerate([('Mode',self.mode),('TCP IP',self.host),('TCP Port',self.tcp_port),('RTU Port',self.com),('Baud',self.baud),('Parity',self.parity),('Stop',self.stop),('Unit ID',self.unit),('Address',self.zero)]): g.addWidget(QLabel(l),0,i); g.addWidget(w,1,i)
        for i,(t,f) in enumerate([('Connect',self.connect),('Disconnect',self.disconnect),('Upload visible',self.upload),('Download selected row',self.download),('Export Excel',self.export_xlsx),('Import Excel',self.import_xlsx)]):
            b=QPushButton(t); b.clicked.connect(f); g.addWidget(b,2,i)
        root.addWidget(box)
        filt=QGroupBox('Search / Filter'); fg=QGridLayout(filt)
        self.search=QLineEdit(); self.search.setPlaceholderText('Search, exact =TTO, address 6005'); self.search.textChanged.connect(self.refresh)
        self.group=QComboBox(); self.group.addItems(['All']+CATEGORIES); self.group.currentTextChanged.connect(self.refresh)
        self.mon=QCheckBox('Monitor only'); self.mon.stateChanged.connect(self.refresh)
        self.keep=QCheckBox('Keep-alive 1s'); self.keep.stateChanged.connect(self.toggle_keep)
        for i,(l,w) in enumerate([('Search',self.search),('Group',self.group),('Filter',self.mon),('Keep alive',self.keep)]): fg.addWidget(QLabel(l),0,i); fg.addWidget(w,1,i)
        root.addWidget(filt)
        self.tabs=QTabWidget(); root.addWidget(self.tabs,1)
        self.table=QTableWidget(); self.tabs.addTab(self.table,'Parameters')
        self.make_diag(); self.make_cmd(); self.make_scope(); self.make_rtu()
        self.logbox=QTextEdit(); self.logbox.setReadOnly(True); self.tabs.addTab(self.logbox,'Event Log')
        self.setStyleSheet('QPushButton{background:#008CD7;color:white;border-radius:8px;min-height:34px;font-weight:bold} QGroupBox{background:white;border:1px solid #ccc;border-radius:10px;margin-top:10px;padding:10px} QMainWindow{background:#EFEFEF}')
    def make_diag(self):
        w=QWidget(); l=QVBoxLayout(w); row=QHBoxLayout()
        for t,codes in [('Read channels',['CRC','CCC','CHCF','FR1','CD1','CD2','CMD','LFR']),('Read communication',['ADD','TTO','M1EC','M1CT','COM1','ETHL','ETHF','TTOB']),('Read state',['ETA','HMIS','RFR','FRH','LFR','LFT','ERRD'])]:
            b=QPushButton(t); b.clicked.connect(lambda _,x=codes:self.read_codes(x)); row.addWidget(b)
        l.addLayout(row); self.diag=QTextEdit(); self.diag.setReadOnly(True); l.addWidget(self.diag); self.tabs.addTab(w,'Communication Diagnostics')
    def make_cmd(self):
        w=QWidget(); l=QVBoxLayout(w); row=QHBoxLayout(); self.cmdlog=QTextEdit('Dry-run only. No CMD@8501 write is performed.'); self.cmdlog.setReadOnly(True)
        for t in ['START','STOP','FAULT RESET']:
            b=QPushButton('Dry Run '+t); b.clicked.connect(lambda _,x=t:self.cmdlog.append(f'[{time.strftime("%H:%M:%S")}] DRY RUN {x}: no CMD write executed.')); row.addWidget(b)
        l.addLayout(row); l.addWidget(self.cmdlog); self.tabs.addTab(w,'Command Dry-Run')
    def make_scope(self):
        w=QWidget(); l=QVBoxLayout(w); row=QHBoxLayout()
        self.scope_interval=QComboBox(); self.scope_interval.addItems(['250','500','1000']); self.scope_window=QComboBox(); self.scope_window.addItems(['30','60','120'])
        row.addWidget(QLabel('Interval ms')); row.addWidget(self.scope_interval); row.addWidget(QLabel('Window s')); row.addWidget(self.scope_window)
        for t,f in [('Start capture',self.start_scope),('Stop capture',self.stop_scope),('Clear',self.clear_scope),('Export CSV',self.export_scope)]:
            b=QPushButton(t); b.clicked.connect(f); row.addWidget(b)
        l.addLayout(row)
        self.plot=pg.PlotWidget(title='XD4000 Oscilloscope / Trend') if pg else QTextEdit('pyqtgraph not installed')
        l.addWidget(self.plot,1); self.tabs.addTab(w,'Oscilloscope / Trend')
    def make_rtu(self):
        w=QWidget(); l=QVBoxLayout(w)
        txt=QTextEdit('RTU RS485 test plan:\n1. Match ADD/Unit ID and serial format.\n2. Start 19200, 8E1 unless drive differs.\n3. Check A/B polarity, common, termination.\n4. Connect RTU and upload =ETA.\n5. Upload =TTO and =LFR.\n6. Test TTO 10 -> 12 -> 10 and LFR low value/readback.\n7. Validate group-wise only.'); txt.setReadOnly(True)
        l.addWidget(txt); self.tabs.addTab(w,'RTU RS485 Test Plan')
    def log(self,m): self.logbox.append(f'[{time.strftime("%H:%M:%S")}] {m}')
    def refresh(self): self.params=self.db.filt(self.search.text() if hasattr(self,'search') else '', self.group.currentText() if hasattr(self,'group') else 'All', self.mon.isChecked() if hasattr(self,'mon') else False); self.populate()
    def populate(self):
        heads=['Code','Group','Subcategory','Name','Address','Type','Scale','Offline','Online','Unit','Access','Protect','Policy']; self.table.blockSignals(True); self.table.setColumnCount(len(heads)); self.table.setHorizontalHeaderLabels(heads); self.table.setRowCount(len(self.params))
        for r,p in enumerate(self.params):
            vals=[p.code,p.group,p.sub,p.name,p.address,p.datatype,p.scale,p.eff,'' if p.online is None else self.fmt(p.online),p.unit,p.access,'Yes' if p.wp else 'No',p.policy]
            for c,v in enumerate(vals):
                it=QTableWidgetItem(str(v)); it.setFlags((it.flags()|Qt.ItemIsEditable) if c==7 and p.access=='RW' and not p.wp else (it.flags()&~Qt.ItemIsEditable)); self.table.setItem(r,c,it)
        self.table.blockSignals(False)
        try: self.table.itemChanged.disconnect()
        except: pass
        self.table.itemChanged.connect(self.edit); self.table.resizeColumnsToContents()
    def fmt(self,v):
        try: return f'{float(v):.3f}'.rstrip('0').rstrip('.')
        except: return str(v)
    def edit(self,item):
        if item.column()!=7: return
        p=self.params[item.row()]
        try:
            v=float(item.text())
            if not(p.min<=v<=p.max): raise ValueError(f'Allowed range {p.min} to {p.max}')
            p.value=v; p.mod=True; self.log(f'Offline changed {p.code}={v}')
        except Exception as e: QMessageBox.warning(self,'Invalid value',str(e))
    def connect(self):
        try:
            if self.mode.currentText()=='TCP': self.gw.connect_tcp(self.host.text().strip(),self.tcp_port.value(),self.unit.value(),self.zero.isChecked())
            else: self.gw.connect_rtu(self.com.text().strip(),int(self.baud.currentText()),self.parity.currentText(),int(self.stop.currentText()),int(self.bits.currentText()),self.unit.value(),self.zero.isChecked())
            self.log(f'Connected {self.gw.mode}')
        except Exception as e: self.log(f'Connection failed: {e}'); QMessageBox.critical(self,'Connection failed',str(e))
    def disconnect(self): self.keep.setChecked(False); self.stop_scope(); self.gw.close(); self.log('Disconnected')
    def upload(self):
        if not self.gw.ok(): QMessageBox.warning(self,'Not connected','Connect first'); return
        ok=fail=0
        for p in self.params:
            try: p.online=self.gw.readp(p); p.value=p.online if not p.mod else p.value; ok+=1; self.log(f'Upload OK {p.code}@{p.address}={self.fmt(p.online)} {p.unit}')
            except Exception as e: fail+=1; self.log(f'Upload failed {p.code}@{p.address}: {e}')
        self.populate(); self.log(f'Upload complete. OK={ok}, Failed={fail}')
    def download(self):
        if not self.gw.ok(): QMessageBox.warning(self,'Not connected','Connect first'); return
        r=self.table.currentRow()
        if r<0: return
        p=self.params[r]
        if p.access!='RW' or p.wp: QMessageBox.warning(self,'Blocked',f'{p.code} is read-only or write-protected'); return
        if QMessageBox.question(self,'Confirm write',f'Write {p.code}@{p.address}={p.eff} {p.unit}?')!=QMessageBox.Yes: return
        try: self.gw.writep(p,p.eff); p.online=self.gw.readp(p); p.value=p.online; p.mod=False; self.populate(); self.log(f'Download/readback OK {p.code}={self.fmt(p.online)}')
        except Exception as e: self.log(f'Download failed {p.code}: {e}')
    def read_codes(self,codes,quiet=False):
        lines=[]
        for code in codes:
            p=self.db.by(code)
            if p:
                try: p.online=self.gw.readp(p); lines.append(f'{code}@{p.address}={self.fmt(p.online)} {p.unit}')
                except Exception as e: lines.append(f'{code} failed: {e}')
        if not quiet: self.diag.append('\n'.join(lines)+'\n')
    def toggle_keep(self): self.keep_timer.start(1000) if self.keep.isChecked() and self.gw.ok() else self.keep_timer.stop()
    def keep_tick(self): self.read_codes(['ETA','RFR','FRH','CRC','CCC','LFR'],quiet=True)
    def start_scope(self):
        if not pg or not self.gw.ok(): return
        self.scope_data={}; self.scope_start=time.time(); self.scope_timer.start(int(self.scope_interval.currentText())); self.log('Scope started')
    def stop_scope(self):
        if self.scope_timer.isActive(): self.scope_timer.stop(); self.log('Scope stopped')
    def clear_scope(self):
        self.scope_data={}
        if pg: self.plot.clear(); self.plot.addLegend()
    def scope_tick(self):
        t=time.time()-(self.scope_start or time.time()); win=float(self.scope_window.currentText()); codes=['RFR','FRH','LFR','ULN','VBUS','THD','LCR']
        for code in codes:
            p=self.db.by(code)
            if p:
                try: v=self.gw.readp(p); self.scope_data.setdefault(code,[]).append((t,v)); self.scope_data[code]=[(x,y) for x,y in self.scope_data[code] if t-x<=win]
                except Exception as e: self.log(f'Scope read failed {code}: {e}')
        if pg:
            self.plot.clear(); self.plot.addLegend(); colors=['#0766F6','#DC272D','#00943D','#FFDD49','#8886FB','#F2AC59','#585B5B']
            for i,(code,pts) in enumerate(self.scope_data.items()):
                if pts: xs,ys=zip(*pts); self.plot.plot(list(xs),list(ys),pen=pg.mkPen(colors[i%len(colors)],width=2),name=code)
    def export_scope(self):
        path,_=QFileDialog.getSaveFileName(self,'Export scope CSV','xd4000_scope.csv','CSV Files (*.csv)')
        if path:
            with open(path,'w',newline='',encoding='utf-8') as f:
                w=csv.writer(f); w.writerow(['signal','time_s','value'])
                for code,pts in self.scope_data.items():
                    for t,v in pts: w.writerow([code,f'{t:.3f}',v])
    def export_xlsx(self):
        if not Workbook: return
        path,_=QFileDialog.getSaveFileName(self,'Export Excel','XD4000_Phase5E_Project.xlsx','Excel Files (*.xlsx)')
        if not path: return
        wb=Workbook(); ws=wb.active; ws.title='Parameters'; ws.append(['code','name','address','datatype','scale','offline_value','online_value','unit','access','write_protect','group','subcategory','policy'])
        for p in self.db.params: ws.append([p.code,p.name,p.address,p.datatype,p.scale,p.eff,'' if p.online is None else p.online,p.unit,p.access,p.wp,p.group,p.sub,p.policy])
        for cell in ws[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='008CD7')
        wb.save(path); self.log(f'Excel exported: {path}')
    def import_xlsx(self):
        if not load_workbook: return
        path,_=QFileDialog.getOpenFileName(self,'Import Excel','','Excel Files (*.xlsx)')
        if not path: return
        wb=load_workbook(path,data_only=True); ws=wb['Parameters']; hdr=[c.value for c in ws[1]]; idx={h:i for i,h in enumerate(hdr)}; n=0
        for row in ws.iter_rows(min_row=2,values_only=True):
            p=self.db.by(str(row[idx.get('code',0)]))
            if p and 'offline_value' in idx and row[idx['offline_value']] not in (None,''):
                try: p.value=float(row[idx['offline_value']]); p.mod=True; n+=1
                except: pass
        self.refresh(); self.log(f'Excel imported. Offline values updated: {n}')
if __name__=='__main__':
    app=QApplication(sys.argv); w=W(); w.show(); sys.exit(app.exec())
